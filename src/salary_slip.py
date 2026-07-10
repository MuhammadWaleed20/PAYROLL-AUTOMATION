import os
from openpyxl import Workbook
from openpyxl.styles import Font


def generate_salary_slips(payroll_df, output_folder):

    os.makedirs(output_folder, exist_ok=True)

    for _, employee in payroll_df.iterrows():

        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Slip"

        # --------------------------
        # Company Information
        # --------------------------

        ws["A1"] = "ABC Technologies Pvt. Ltd."
        ws["A1"].font = Font(size=16, bold=True)

        ws["A2"] = "Monthly Salary Slip"
        ws["A2"].font = Font(size=13, bold=True)

        # --------------------------
        # Employee Information
        # --------------------------

        ws["A4"] = "Employee ID"
        ws["B4"] = employee["Employee ID"]

        ws["A5"] = "Name"
        ws["B5"] = employee["Name"]

        ws["A6"] = "Department"
        ws["B6"] = employee["Department"]

        # --------------------------
        # Payroll Information
        # --------------------------

        ws["A8"] = "Basic Salary"
        ws["B8"] = employee["Basic Salary"]

        ws["A9"] = "Overtime Pay"
        ws["B9"] = employee["Overtime Pay"]

        ws["A10"] = "Bonus"
        ws["B10"] = employee["Bonus"]

        ws["A11"] = "Leave Deduction"
        ws["B11"] = employee["Leave Deduction"]

        ws["A12"] = "Tax Amount"
        ws["B12"] = employee["Tax Amount"]

        ws["A13"] = "Net Salary"
        ws["B13"] = employee["Net Salary"]

        ws["A13"].font = Font(bold=True)
        ws["B13"].font = Font(bold=True)

        filename = f"{employee['Employee ID']}_{employee['Name']}.xlsx"

        filepath = os.path.join(output_folder, filename)

        wb.save(filepath)

    print("✅ Salary slips generated successfully.")